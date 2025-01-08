from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import datetime
import openpyxl

from .models import ManufacturingOrder, Material

def upload_page(request):
    return render(request, 'api/upload.html')

@csrf_exempt
def upload_file(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Example of extracting data
        order_number = ws.cell(row=1, column=18).value  # 假設製令單號位於 (1, 18)

        # 檢查是否已存在該製令單號
        if ManufacturingOrder.objects.filter(order_number=order_number).exists():
            return JsonResponse({'status': 'error', 'message': f'製令單號 {order_number} 已存在'}, status=400)

        # 新建製令單
        order = ManufacturingOrder.objects.create(order_number=order_number)

        # 插入材料資料 (假設從第 14 行開始是材料)
        for row in ws.iter_rows(min_row=14, max_col=16, values_only=True):
            if row[0]:  # 確認有數據
                Material.objects.create(
                    order=order,
                    part_name=row[0],
                    material_name=row[2],
                    quantity=row[16],
                    unit=row[15]
                )
        return JsonResponse({'status': 'success', 'message': '文件上傳成功'})
    return JsonResponse({'status': 'error', 'message': '無效的請求'}, status=400)
